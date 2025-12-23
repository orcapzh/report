"""
合并所有送货单的出货品类及数量
"""
import pandas as pd
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

def extract_data_from_excel(file_path):
    """从单个Excel文件提取数据"""
    print(f"正在处理: {file_path}")

    try:
        # 读取Excel文件，不设置header
        df = pd.read_excel(file_path, sheet_name=0, header=None)

        # 提取客户名称和日期
        customer_name = None
        date = None
        if len(df) > 6:
            # 第6行包含客户名称和日期
            row_6 = df.iloc[6]
            if pd.notna(row_6[2]):
                customer_name = str(row_6[2]).strip()
            if pd.notna(row_6[8]):
                date = row_6[8]

        # 数据从第10行开始，直到遇到"合计金额"
        data_rows = []
        for idx in range(10, len(df)):
            row = df.iloc[idx]

            # 检查是否到达合计行
            if pd.notna(row[1]) and '合计' in str(row[1]):
                break

            # 提取货名、规格、数量、单位、单价、金额
            product_name = row[1] if pd.notna(row[1]) else None
            spec = row[3] if pd.notna(row[3]) else None
            quantity = row[5] if pd.notna(row[5]) else None
            unit = row[6] if pd.notna(row[6]) else None
            unit_price = row[7] if pd.notna(row[7]) else None
            amount = row[8] if pd.notna(row[8]) else None

            # 只保留有效的数据行（至少有货名和数量）
            if product_name and quantity:
                # 清理货名中的换行符
                product_name = str(product_name).replace('\n', ' ').strip()
                if spec:
                    spec = str(spec).strip()
                if unit:
                    unit = str(unit).strip()

                data_rows.append({
                    '货名': product_name,
                    '规格': spec if spec else '',
                    '数量': float(quantity),
                    '单位': unit if unit else '',
                    '单价': float(unit_price) if unit_price else 0,
                    '金额': float(amount) if amount else 0,
                    '客户': customer_name if customer_name else '',
                    '日期': date if date else '',
                    '文件': os.path.basename(file_path)
                })

        return data_rows

    except Exception as e:
        print(f"  处理 {file_path} 时出错: {e}")
        import traceback
        traceback.print_exc()
        return []

def merge_delivery_orders(raw_data_dir='raw-data', output_file='output/merged_delivery_orders.xlsx'):
    """合并所有送货单"""

    # 创建输出目录
    output_path = Path(output_file).parent
    output_path.mkdir(parents=True, exist_ok=True)

    # 查找所有Excel文件
    raw_data_path = Path(raw_data_dir)
    excel_files = list(raw_data_path.glob('**/*.xls')) + list(raw_data_path.glob('**/*.xlsx'))

    print(f"找到 {len(excel_files)} 个Excel文件")
    print()

    # 提取所有数据
    all_data = []
    for file_path in excel_files:
        data = extract_data_from_excel(file_path)
        all_data.extend(data)

    print(f"\n共提取 {len(all_data)} 条数据记录")

    if not all_data:
        print("没有找到任何数据")
        return

    # 转换为DataFrame
    df_all = pd.DataFrame(all_data)

    # 按货名和规格分组，汇总数量和金额
    print("\n正在合并相同的货名和规格...")
    df_summary = df_all.groupby(['货名', '规格', '单位']).agg({
        '数量': 'sum',
        '金额': 'sum',
        '客户': lambda x: ', '.join(sorted(set(filter(None, x)))),
        '文件': lambda x: ', '.join(sorted(set(x)))
    }).reset_index()

    # 计算平均单价
    df_summary['平均单价'] = df_summary['金额'] / df_summary['数量']
    df_summary['平均单价'] = df_summary['平均单价'].round(2)

    # 调整列顺序
    df_summary = df_summary[['货名', '规格', '单位', '数量', '平均单价', '金额', '客户', '文件']]

    # 按金额降序排列
    df_summary = df_summary.sort_values('金额', ascending=False)

    # 生成透视分析数据
    print(f"\n正在生成透视分析...")

    # 1. 按客户汇总
    df_by_customer = df_all.groupby('客户').agg({
        '数量': 'sum',
        '金额': 'sum',
        '货名': 'count'
    }).rename(columns={'货名': '订单数'}).reset_index()
    df_by_customer['平均单价'] = (df_by_customer['金额'] / df_by_customer['数量']).round(2)
    df_by_customer = df_by_customer.sort_values('金额', ascending=False)
    df_by_customer = df_by_customer[['客户', '订单数', '数量', '金额', '平均单价']]

    # 2. 按品名汇总
    df_by_product = df_all.groupby(['货名', '规格', '单位']).agg({
        '数量': 'sum',
        '金额': 'sum',
        '客户': lambda x: ', '.join(sorted(set(filter(None, x))))
    }).reset_index()
    df_by_product['平均单价'] = (df_by_product['金额'] / df_by_product['数量']).round(2)
    df_by_product = df_by_product.sort_values('金额', ascending=False)
    df_by_product = df_by_product[['货名', '规格', '单位', '数量', '平均单价', '金额', '客户']]

    # 3. 按月份汇总
    df_all['月份'] = pd.to_datetime(df_all['日期']).dt.to_period('M').astype(str)
    df_by_month = df_all.groupby('月份').agg({
        '数量': 'sum',
        '金额': 'sum',
        '货名': 'count',
        '客户': lambda x: len(set(filter(None, x)))
    }).rename(columns={'货名': '订单数', '客户': '客户数'}).reset_index()
    df_by_month['平均订单金额'] = (df_by_month['金额'] / df_by_month['订单数']).round(2)
    df_by_month = df_by_month[['月份', '订单数', '客户数', '数量', '金额', '平均订单金额']]

    # 4. 客户月度交叉分析
    df_customer_month = df_all.groupby(['客户', '月份']).agg({
        '数量': 'sum',
        '金额': 'sum',
        '货名': 'count'
    }).rename(columns={'货名': '订单数'}).reset_index()
    df_customer_month = df_customer_month.sort_values(['客户', '月份'])
    df_customer_month = df_customer_month[['客户', '月份', '订单数', '数量', '金额']]

    # 保存详细数据和汇总数据到Excel
    print(f"\n正在保存到 {output_file}...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 汇总数据
        df_summary.to_excel(writer, sheet_name='汇总', index=False)

        # 详细数据
        df_all_sorted = df_all.sort_values(['货名', '规格', '日期'])
        df_all_sorted.to_excel(writer, sheet_name='详细数据', index=False)

        # 透视分析数据
        df_by_customer.to_excel(writer, sheet_name='按客户分析', index=False)
        df_by_product.to_excel(writer, sheet_name='按产品分析', index=False)
        df_by_month.to_excel(writer, sheet_name='按月份分析', index=False)
        df_customer_month.to_excel(writer, sheet_name='客户月度分析', index=False)

    print(f"\n合并完成！")
    print(f"汇总数据共 {len(df_summary)} 种品类")
    print(f"\n透视分析:")
    print(f"- 客户数: {len(df_by_customer)}")
    print(f"- 产品数: {len(df_by_product)}")
    print(f"- 月份数: {len(df_by_month)}")
    print(f"\n汇总预览:")
    print(df_summary.to_string())

    return df_summary

def amount_to_chinese(amount):
    """将金额转换为中文大写"""
    chinese_numbers = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_units = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']

    # 处理小数部分
    amount_str = f"{amount:.2f}"
    integer_part, decimal_part = amount_str.split('.')

    # 转换整数部分
    result = ''
    integer_part = integer_part[::-1]  # 反转字符串

    for i, digit in enumerate(integer_part):
        digit_value = int(digit)
        if digit_value != 0:
            result = chinese_numbers[digit_value] + chinese_units[i] + result
        else:
            if result and result[0] != '零':
                result = '零' + result

    # 清理多余的零
    while '零零' in result:
        result = result.replace('零零', '零')

    if result.endswith('零'):
        result = result[:-1]

    if not result:
        result = '零'

    result += '元'

    # 处理角分
    jiao = int(decimal_part[0])
    fen = int(decimal_part[1])

    if jiao == 0 and fen == 0:
        result += '整'
    else:
        if jiao != 0:
            result += chinese_numbers[jiao] + '角'
        if fen != 0:
            result += chinese_numbers[fen] + '分'

    return result

def create_statement(df_all, customer_name, year_month, output_file='statement.xlsx',
                     company_name='百惠行对账单',
                     address='东莞市黄江镇华南塑胶城区132号',
                     phone='(0769) 83631717',
                     fax='83637787'):
    """生成对账单"""

    print(f"\n正在生成对账单到 {output_file}...")

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '对账单'

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # 标题行 (第1行)
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = company_name
    title_cell.font = Font(name='宋体', size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 地址行 (第2行)
    ws.merge_cells('A2:G2')
    address_cell = ws['A2']
    address_cell.value = f'地址：{address}'
    address_cell.font = Font(name='宋体', size=10)
    address_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 联系方式行 (第3行)
    ws.merge_cells('A3:G3')
    contact_cell = ws['A3']
    contact_cell.value = f'电话：{phone}    传真：{fax}'
    contact_cell.font = Font(name='宋体', size=10)
    contact_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 客户和日期信息 (第4行)
    ws.merge_cells('A4:B4')
    ws['A4'] = f'客户：{customer_name}'
    ws['A4'].alignment = Alignment(horizontal='left')
    ws.merge_cells('C4:E4')
    ws['C4'] = f'{year_month}对账单'
    ws['C4'].alignment = Alignment(horizontal='center')

    # 表头 (第5行)
    headers = ['送货日期', '品名规格', '单位', '数量', '单价', '金额', '备注']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(name='宋体', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 数据行
    row_num = 6
    total_amount = 0

    # 按日期排序数据
    df_sorted = df_all.sort_values('日期')

    for _, row_data in df_sorted.iterrows():
        # 格式化日期 - 只显示日期部分
        date_obj = row_data['日期']
        if hasattr(date_obj, 'strftime'):
            date_str = date_obj.strftime('%Y-%m-%d')
        else:
            date_str = str(date_obj).split('T')[0] if 'T' in str(date_obj) else str(date_obj)

        ws.cell(row=row_num, column=1, value=date_str)
        ws.cell(row=row_num, column=2, value=f"{row_data['货名']} {row_data['规格']}")
        ws.cell(row=row_num, column=3, value=row_data['单位'])
        ws.cell(row=row_num, column=4, value=row_data['数量'])
        ws.cell(row=row_num, column=5, value=row_data['单价'])
        ws.cell(row=row_num, column=6, value=row_data['金额'])
        ws.cell(row=row_num, column=7, value='')  # 备注

        total_amount += row_data['金额']

        # 设置对齐和边框
        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            # 品名规格列设置自动换行
            if col_num == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = Font(name='宋体', size=10)

        row_num += 1

    # 为表格添加边框（包括表头）
    for col_num in range(1, 8):
        ws.cell(row=5, column=col_num).border = thin_border

    # 合计行（空几行后）
    summary_row = row_num + 2
    ws.merge_cells(f'A{summary_row}:C{summary_row}')

    # 中文大写金额
    chinese_amount = amount_to_chinese(total_amount)
    ws[f'A{summary_row}'] = f'合计人民币大写：{chinese_amount}'
    ws[f'A{summary_row}'].font = Font(name='宋体', size=11)

    # 小写金额
    ws.merge_cells(f'D{summary_row}:G{summary_row}')
    ws[f'D{summary_row}'] = f'人民币小写：{total_amount:.2f}元'
    ws[f'D{summary_row}'].font = Font(name='宋体', size=11)
    ws[f'D{summary_row}'].alignment = Alignment(horizontal='right')

    # 设置打印选项
    ws.page_setup.paperSize = 9  # A4纸
    ws.page_setup.orientation = 'portrait'  # 纵向
    ws.page_setup.fitToWidth = 1  # 适配宽度为1页
    ws.page_setup.fitToHeight = 0  # 高度不限制（自动）

    # 设置页边距（单位：英寸）
    ws.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3
    )

    # 设置打印标题（每页都显示表头）
    ws.print_title_rows = '1:5'  # 第1到第5行作为打印标题

    # 保存文件
    wb.save(output_file)
    print(f"对账单已生成: {output_file}")
    print(f"总金额: {total_amount:.2f}元 ({chinese_amount})")

if __name__ == '__main__':
    # 合并送货单
    df_summary = merge_delivery_orders()

    # 读取详细数据用于生成对账单
    df_all = pd.read_excel('output/merged_delivery_orders.xlsx', sheet_name='详细数据')

    # 转换日期列为datetime类型
    df_all['日期'] = pd.to_datetime(df_all['日期'])

    # 提取年月
    df_all['年月'] = df_all['日期'].dt.to_period('M')

    # 按客户和年月分组
    grouped = df_all.groupby(['客户', '年月'])

    print(f"\n\n===== 开始生成对账单 =====")
    print(f"共有 {len(grouped)} 个客户月份组合\n")

    # 创建输出目录
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)

    # 为每个客户的每个月生成对账单
    skipped_count = 0
    generated_count = 0

    for (customer, year_month), group_data in grouped:
        # 创建客户文件夹
        customer_dir = output_dir / customer
        customer_dir.mkdir(exist_ok=True)

        # 生成文件名
        output_file = customer_dir / f'statement_{customer}_{year_month}.xlsx'

        # 检查文件是否已存在
        if output_file.exists():
            print(f"\n对账单已存在，跳过: {output_file}")
            skipped_count += 1
            continue

        # 格式化年月显示
        year_month_str = f'{year_month.year}年{year_month.month}月'

        # 生成对账单
        create_statement(
            group_data,
            customer_name=customer,
            year_month=year_month_str,
            output_file=str(output_file)
        )
        generated_count += 1

    print(f"\n\n===== 所有对账单生成完成 =====")
    print(f"新生成: {generated_count} 个对账单")
    print(f"已跳过: {skipped_count} 个对账单")
    print(f"文件保存位置: output/ 文件夹")
